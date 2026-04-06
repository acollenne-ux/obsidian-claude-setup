"""
email_trigger.py — Bridge mobile via Gmail.

Fonctionnement :
  1. Surveille la boite Gmail toutes les 30 secondes
  2. Detecte les emails dont le sujet contient le mot "claude" (insensible a la casse)
  3. Construit le prompt a partir du corps + pieces jointes (PDF, texte, images, Word, CSV...)
  4. Lance Claude CLI avec la demande complete
  5. Envoie la reponse par email + PDF

Depuis le telephone :
  - Envoyer un email a acollenne@gmail.com
  - Sujet : "Claude analyse Microsoft" ou "Rapport claude trading" ou "claude: question"
  - Corps : le texte complet de ta demande (peut etre long)
  - Pieces jointes : images, PDF, Word, CSV, TXT... tout est supporte

Lancement : python email_trigger.py
"""
import imaplib, smtplib, email, json, subprocess, sys, os, time, re, logging, tempfile, threading
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from pathlib import Path

# Config
CONFIG_PATH  = Path(__file__).parent / "email_config.json"
REPORTS_DIR  = Path(__file__).parent / "reports"
ATTACH_DIR   = Path(__file__).parent / "attachments"
DISPATCH_FILE = Path(__file__).parent / "dispatch_inbox.json"
CLAUDE_CMD   = r"C:\Users\Alexandre collenne\.claude\tools\run_claude.bat"
POLL_SECS    = 10
TRIGGER_WORD = "claude travaille"   # sujet doit contenir cette phrase exacte (insensible a la casse)
# Mode dispatch : True = ecrire dans dispatch_inbox.json pour pickup par Claude Code Desktop
# False = lancer Claude CLI directement (ancien comportement)
DISPATCH_MODE = True

# Instruction systeme injectee via --append-system-prompt (autorite systeme, pas injection utilisateur)
# Single-line obligatoire : cmd.exe tronque les args au premier \n
SYSTEM_INSTRUCTION = (
    "Tu es l assistant Claude Code d Alexandre Collenne. "
    "Pour TOUTES les demandes recues via ce bridge email, tu DOIS invoquer "
    "le skill deep-research (Skill tool avec skill=deep-research) AVANT de formuler ta reponse. "
    "Ce skill orchestre plusieurs IAs, recherche sur internet, et genere un PDF envoye a acollenne@gmail.com. "
    "Cela s applique meme pour les demandes simples."
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(Path(__file__).parent / "email_trigger.log", encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)


# ─── Extraction de texte depuis pieces jointes ──────────────────────────────

def extract_pdf_text(filepath):
    """Extrait le texte d'un PDF (pdfplumber > pypdf > fallback)."""
    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
        text = "\n".join(pages).strip()
        if text:
            return text
    except ImportError:
        pass
    except Exception as e:
        log.warning(f"pdfplumber echec: {e}")

    try:
        import pypdf
        reader = pypdf.PdfReader(filepath)
        pages = [p.extract_text() or "" for p in reader.pages]
        return "\n".join(pages).strip()
    except ImportError:
        pass
    except Exception as e:
        log.warning(f"pypdf echec: {e}")

    return f"[PDF joint: {Path(filepath).name} — impossible d'extraire le texte, installe pdfplumber: pip install pdfplumber]"


def extract_docx_text(filepath):
    """Extrait le texte d'un fichier Word .docx."""
    try:
        from docx import Document
        doc = Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except ImportError:
        return f"[Word joint: {Path(filepath).name} — installe python-docx: pip install python-docx]"
    except Exception as e:
        return f"[Word joint: {Path(filepath).name} — erreur: {e}]"


def extract_xlsx_text(filepath):
    """Extrait les donnees d'un fichier Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"[Feuille: {sheet}]")
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(c) if c is not None else "" for c in row)
                if row_str.strip(" |"):
                    lines.append(row_str)
        return "\n".join(lines)
    except ImportError:
        return f"[Excel joint: {Path(filepath).name} — installe openpyxl: pip install openpyxl]"
    except Exception as e:
        return f"[Excel joint: {Path(filepath).name} — erreur: {e}]"


def process_attachment(part, filename):
    """
    Sauvegarde la piece jointe et extrait son contenu textuel si possible.
    Retourne (chemin_fichier, texte_extrait_ou_None, est_image).
    """
    ATTACH_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[^\w.\-]', '_', filename)
    filepath  = ATTACH_DIR / safe_name

    data = part.get_payload(decode=True)
    with open(filepath, "wb") as f:
        f.write(data)

    ext = Path(filename).suffix.lower()

    # Fichiers texte simples
    if ext in (".txt", ".md", ".csv", ".log", ".json", ".xml", ".html", ".htm", ".py", ".js", ".ts"):
        try:
            text = data.decode("utf-8", errors="replace")
            return str(filepath), f"[Fichier joint: {filename}]\n{text}", False
        except Exception:
            return str(filepath), f"[Fichier joint: {filename} — impossible de lire]", False

    # PDF
    if ext == ".pdf":
        text = extract_pdf_text(str(filepath))
        return str(filepath), f"[PDF joint: {filename}]\n{text}", False

    # Word
    if ext in (".docx", ".doc"):
        text = extract_docx_text(str(filepath))
        return str(filepath), f"[Word joint: {filename}]\n{text}", False

    # Excel
    if ext in (".xlsx", ".xls"):
        text = extract_xlsx_text(str(filepath))
        return str(filepath), f"[Excel joint: {filename}]\n{text}", False

    # Images
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".heic"):
        return str(filepath), None, True

    # Fichier inconnu : juste mentionner
    return str(filepath), f"[Piece jointe: {filename} (type non supporte: {ext})]", False


# ─── Recuperation des emails ────────────────────────────────────────────────

def decode_subject(raw_subject):
    """Decode proprement un sujet d'email (gere UTF-8, latin-1, etc.)."""
    parts = email.header.decode_header(raw_subject or "")
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(part)
    return " ".join(decoded).strip()


def fetch_pending(cfg):
    """
    Retourne la liste des emails non lus a traiter.
    Chaque element : (uid, sujet, corps_texte, liste_pieces_jointes, liste_images)
    """
    pending = []
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        mail.login(cfg["sender_email"], cfg["gmail_app_password"])
        mail.select("inbox")

        # Cherche les emails non lus envoyes par le proprietaire
        _, data = mail.search(None, '(UNSEEN FROM "acollenne@gmail.com")')
        uids = data[0].split()
        log.info(f"{len(uids)} email(s) non lu(s) a verifier")

        for uid in uids:
            _, msg_data = mail.fetch(uid, "(RFC822)")
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            subject = decode_subject(msg["Subject"])

            # Filtre 1 : ignorer les reponses et transferts (Re:, Fwd:)
            subject_stripped = subject.lower().lstrip()
            if subject_stripped.startswith("re:") or subject_stripped.startswith("fwd:"):
                log.info(f"Ignore (reponse/transfert): {subject}")
                mail.store(uid, "+FLAGS", "\\Seen")
                continue

            # Filtre 2 : ignorer si header de reponse present (In-Reply-To / References)
            if msg.get("In-Reply-To") or msg.get("References"):
                log.info(f"Ignore (header reponse detecte): {subject}")
                mail.store(uid, "+FLAGS", "\\Seen")
                continue

            # Filtre 3 : le sujet doit contenir "claude travaille"
            if TRIGGER_WORD not in subject.lower():
                log.info(f"Ignore (pas de '{TRIGGER_WORD}' dans le sujet): {subject}")
                continue

            # Extraire corps + pieces jointes
            body_text  = ""
            body_html  = ""
            attach_texts = []
            image_paths  = []

            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    content_disp = str(part.get("Content-Disposition") or "")

                    # Piece jointe
                    if "attachment" in content_disp or part.get_filename():
                        filename = part.get_filename()
                        if filename:
                            filename = decode_subject(filename)
                            log.info(f"Piece jointe detectee: {filename}")
                            filepath, text_content, is_image = process_attachment(part, filename)
                            if is_image:
                                image_paths.append(filepath)
                            elif text_content:
                                attach_texts.append(text_content)
                    # Corps texte
                    elif content_type == "text/plain" and not body_text:
                        body_text = part.get_payload(decode=True).decode("utf-8", errors="replace")
                    elif content_type == "text/html" and not body_html:
                        body_html = part.get_payload(decode=True).decode("utf-8", errors="replace")
            else:
                # Email simple
                payload = msg.get_payload(decode=True)
                if payload:
                    body_text = payload.decode("utf-8", errors="replace")

            # Si pas de texte brut, extraire le texte du HTML
            if not body_text and body_html:
                # Nettoyage HTML basique
                body_text = re.sub(r'<[^>]+>', ' ', body_html)
                body_text = re.sub(r'\s+', ' ', body_text).strip()

            # Marquer comme lu
            body_stripped = body_text.strip()
            mail.store(uid, "+FLAGS", "\\Seen")
            pending.append((uid, subject, body_stripped, attach_texts, image_paths))
            log.info(f"Email a traiter: '{subject}' | corps={len(body_stripped)} chars | "
                     f"PJ texte={len(attach_texts)} | images={len(image_paths)}")
            if body_stripped:
                log.debug(f"Corps (debut): {body_stripped[:200]}")

        mail.logout()
    except Exception as e:
        log.error(f"Erreur IMAP: {e}")
    return pending


# ─── Construction du prompt ─────────────────────────────────────────────────

def build_prompt(subject, body, attach_texts, image_paths):
    """
    Construit la demande utilisateur pour Claude.
    L instruction deep-research est passee separement via --append-system-prompt
    (evite la detection d injection de prompt).
    """
    clean_subject = re.sub(r'(?i)claude\s+travaille\s*[-:,]?\s*', '', subject).strip()

    parts = []

    # Demande principale : sujet + corps
    if clean_subject and body:
        parts.append(f"Objet : {clean_subject}\n\n{body}")
    elif body:
        parts.append(body)
    elif clean_subject:
        parts.append(clean_subject)
    else:
        parts.append(subject)

    # Pieces jointes textuelles
    if attach_texts:
        parts.append("\n\n--- PIECES JOINTES ---")
        for t in attach_texts:
            parts.append(f"\n{t}")

    # Images
    if image_paths:
        parts.append(f"\n\n--- IMAGES JOINTES ({len(image_paths)}) ---")
        for p in image_paths:
            parts.append(f"Image disponible : {p}")
        parts.append("(Note: les images ont ete sauvegardees localement. "
                      "Mentionne-les dans ton analyse si pertinent.)")

    return "\n".join(parts)


# ─── Execution Claude ────────────────────────────────────────────────────────

def run_claude(prompt):
    """Execute Claude CLI et retourne la reponse.

    IMPORTANT : le prompt est passe via stdin (-p -) et non comme argument CLI.
    Raison : Windows cmd.exe tronque les arguments au premier retour a la ligne,
    ce qui faisait que Claude ne recevait que l'instruction systeme sans la demande.
    Solution : subprocess passe le prompt via stdin (input=prompt), sans newlines
    dans les arguments de la ligne de commande.
    """
    log.info(f"Lancement Claude ({len(prompt)} chars): {prompt[:120]}...")
    log.debug(f"Prompt complet:\n{prompt}")
    try:
        # -p - : dit a Claude de lire le prompt depuis stdin
        # input=prompt : Python passe le prompt via le pipe stdin
        # shell=True requis pour executer les fichiers .bat Windows
        # --append-system-prompt : instruction deep-research en tant que system prompt
        # (autorite systeme, evite la detection d injection dans le message utilisateur)
        # input=prompt : demande utilisateur via stdin (supporte les newlines, pas de troncature cmd.exe)
        result = subprocess.run(
            [CLAUDE_CMD, "-p",
             "--append-system-prompt", SYSTEM_INSTRUCTION,
             "--dangerously-skip-permissions"],
            input=prompt,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=600,
            cwd=r"C:\Users\Alexandre collenne",
            shell=True,
        )
        response = result.stdout.strip()
        if result.returncode != 0:
            log.warning(f"Claude rc={result.returncode} | stderr: {result.stderr[:400]}")
        if not response:
            response = result.stderr.strip()
        log.info(f"Reponse Claude: {len(response)} caracteres")
        return response or "Pas de reponse obtenue."
    except subprocess.TimeoutExpired:
        return "Timeout : la demande a pris trop de temps (>10 min). Essaie de simplifier la demande."
    except Exception as e:
        log.error(f"Erreur run_claude: {e}")
        return f"Erreur Claude: {e}"


# ─── PDF + email ─────────────────────────────────────────────────────────────

def generate_pdf(subject, content):
    """Genere un PDF via send_report.py et retourne le chemin.
    Le sujet de l'email de reponse NE doit PAS contenir 'claude travaille'
    pour eviter que le bridge ne retraite la reponse comme une nouvelle demande.
    """
    REPORTS_DIR.mkdir(exist_ok=True)
    # Retirer "claude travaille" du sujet de reponse
    clean = re.sub(r'(?i)claude\s+travaille\s*[-:,]?\s*', '', subject).strip()
    report_subject = f"Résultat — {clean}" if clean else "Résultat Claude"
    try:
        result = subprocess.run(
            [sys.executable,
             str(Path(__file__).parent / "send_report.py"),
             report_subject, content, "acollenne@gmail.com"],
            capture_output=True, text=True, timeout=60,
            encoding="utf-8", errors="replace"
        )
        match = re.search(r"PDF : (.+\.pdf)", result.stdout)
        if match:
            return match.group(1).strip()
    except Exception as e:
        log.error(f"Erreur PDF: {e}")
    return None


def send_response(cfg, original_subject, response_text, pdf_path=None, image_paths=None):
    """Envoie la reponse par email avec corps HTML + PDF en piece jointe."""
    sender = cfg["sender_email"]
    pwd    = cfg["gmail_app_password"]

    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = sender
    msg["Subject"] = f"Re: {original_subject}"

    # Corps HTML
    html_body = response_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    html_body = html_body.replace("\n", "<br>")

    nb_images = len(image_paths or [])
    images_note = (f'<p style="font-size:11px;color:#e57;margin-top:8px">'
                   f'Note: {nb_images} image(s) jointe(s) recue(s) — '
                   f'traitees localement.</p>') if nb_images else ""

    html = f"""<html><body style="font-family:Arial;max-width:800px;margin:auto;color:#333">
    <div style="background:#0a2864;padding:16px;border-radius:6px 6px 0 0">
      <h2 style="color:white;margin:0;font-size:16px">Claude Code — Reponse</h2>
      <p style="color:#aac4ff;margin:4px 0 0;font-size:11px">{datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
    </div>
    <div style="background:#f8f9fc;padding:20px;border:1px solid #ddd;border-radius:0 0 6px 6px">
      <p style="font-size:12px;color:#666;margin-bottom:16px">Demande : <strong>{original_subject}</strong></p>
      {images_note}
      <div style="font-size:13px;line-height:1.7">{html_body}</div>
      {'<p style="font-size:11px;color:#aaa;margin-top:20px">Rapport PDF en piece jointe.</p>' if pdf_path else ''}
    </div></body></html>"""

    msg.attach(MIMEText(html, "html"))

    # Attacher le PDF
    if pdf_path and os.path.exists(pdf_path):
        with open(pdf_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{Path(pdf_path).name}"')
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
            srv.login(sender, pwd)
            srv.send_message(msg)
        log.info(f"Reponse envoyee: {original_subject}")
    except Exception as e:
        log.error(f"Erreur envoi email: {e}")


# ─── Dispatch vers Claude Code Desktop ──────────────────────────────────────

def write_dispatch(subject, prompt, image_paths):
    """Ecrit la requete dans dispatch_inbox.json pour pickup par Claude Code Desktop."""
    if DISPATCH_FILE.exists():
        with open(DISPATCH_FILE, 'r', encoding='utf-8') as f:
            try:
                inbox = json.load(f)
            except json.JSONDecodeError:
                inbox = []
    else:
        inbox = []

    clean_subject = re.sub(r'(?i)claude\s+travaille\s*[-:,]?\s*', '', subject).strip()
    inbox.append({
        "id": datetime.now().strftime("%Y%m%d_%H%M%S"),
        "timestamp": datetime.now().isoformat(),
        "subject": clean_subject or subject,
        "prompt": prompt,
        "images": image_paths or [],
        "status": "pending"
    })

    with open(DISPATCH_FILE, 'w', encoding='utf-8') as f:
        json.dump(inbox, f, indent=2, ensure_ascii=False)

    log.info(f"Dispatch: requete '{clean_subject}' ajoutee a {DISPATCH_FILE}")


# ─── Pipeline principal ───────────────────────────────────────────────────────

def process_email(cfg, subject, body, attach_texts, image_paths):
    """Pipeline : dispatch vers Desktop (DISPATCH_MODE=True) ou Claude CLI direct (False)."""
    log.info(f"=== Traitement: {subject} ===")
    prompt = build_prompt(subject, body, attach_texts, image_paths)

    if DISPATCH_MODE:
        write_dispatch(subject, prompt, image_paths)
    else:
        response = run_claude(prompt)
        generate_pdf(subject, response)

    log.info(f"=== Termine: {subject} ===")


def main():
    cfg = load_cfg()
    log.info("=" * 55)
    log.info("  Claude Email Bridge - Demarre")
    log.info("=" * 55)
    log.info(f"Surveillance de {cfg['sender_email']} toutes les {POLL_SECS}s (polling rapide)")
    log.info(f"Phrase declencheuse dans le sujet : '{TRIGGER_WORD}'")
    log.info("")
    log.info("Exemples de sujets valides (doivent contenir 'claude travaille') :")
    log.info("  'claude travaille'")
    log.info("  'claude travaille analyse Microsoft'")
    log.info("  'claude travaille - voir document joint'")
    log.info("")
    log.info("Pieces jointes supportees : PDF, Word, Excel, TXT, CSV, images")
    log.info("=" * 55)

    # Suivi des UIDs en cours de traitement pour eviter les doublons
    processing_uids = set()
    lock = threading.Lock()

    def handle_email(uid, subject, body, attach_texts, image_paths):
        try:
            process_email(cfg, subject, body, attach_texts, image_paths)
        finally:
            with lock:
                processing_uids.discard(uid)

    while True:
        try:
            pending = fetch_pending(cfg)
            for uid, subject, body, attach_texts, image_paths in pending:
                uid_str = uid.decode() if isinstance(uid, bytes) else str(uid)
                with lock:
                    if uid_str in processing_uids:
                        log.info(f"Deja en cours de traitement, ignore : {subject}")
                        continue
                    processing_uids.add(uid_str)
                # Lancer dans un thread separe : le polling continue pendant le traitement
                t = threading.Thread(
                    target=handle_email,
                    args=(uid_str, subject, body, attach_texts, image_paths),
                    daemon=True
                )
                t.start()
                log.info(f"Thread lance pour : {subject} (threads actifs: {threading.active_count()})")
        except KeyboardInterrupt:
            log.info("Arret du bridge email.")
            break
        except Exception as e:
            log.error(f"Erreur boucle principale: {e}")
        time.sleep(POLL_SECS)


def load_cfg():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


if __name__ == "__main__":
    main()
